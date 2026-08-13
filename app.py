from __future__ import annotations

import html
import io
import json
from pathlib import Path

import streamlit as st
import streamlit.components.v1 as components
from openpyxl import load_workbook

APP_DIR = Path(__file__).resolve().parent
DATA_FILENAME = "AUiX_Expertise_Map_Data.xlsx"
DEFAULT_DATA = next(
    (candidate for candidate in (APP_DIR / DATA_FILENAME, APP_DIR / "data" / DATA_FILENAME) if candidate.exists()),
    APP_DIR / DATA_FILENAME,
)

TYPE_ORDER = [
    "Air University",
    "Academia",
    "Industry",
    "Military & Government",
]

TYPE_COLORS = {
    "Air University": "#C94242",
    "Academia": "#3C78D8",
    "Industry": "#D4A62A",
    "Military & Government": "#3E8E5B",
}


def _rows_as_dicts(ws):
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None for v in values):
            continue
        rows.append({headers[i]: values[i] for i in range(len(headers))})
    return rows


@st.cache_data(show_spinner=False)
def load_network_data(file_bytes: bytes | None = None) -> dict:
    """Load the normalized AUiX workbook and return JSON-ready network data."""
    if file_bytes is None:
        wb = load_workbook(DEFAULT_DATA, data_only=True, read_only=True)
    else:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)

    required = {"Stakeholders", "Expertise_Links", "Expertise"}
    missing = sorted(required.difference(wb.sheetnames))
    if missing:
        raise ValueError("Workbook is missing required sheet(s): " + ", ".join(missing))

    stakeholder_rows = _rows_as_dicts(wb["Stakeholders"])
    link_rows = _rows_as_dicts(wb["Expertise_Links"])
    expertise_rows = _rows_as_dicts(wb["Expertise"])

    organizations = {}
    for row in stakeholder_rows:
        name = str(row.get("Organization") or "").strip()
        if not name:
            continue
        org_type = str(row.get("Organization Type") or "Unknown").strip()
        score = row.get("Engagement Score")
        try:
            score = float(score or 0)
        except (TypeError, ValueError):
            score = 0.0
        areas = [x.strip() for x in str(row.get("Expertise Areas") or "").split(";") if x.strip()]
        organizations[name] = {
            "name": name,
            "type": org_type,
            "score": score,
            "color": TYPE_COLORS.get(org_type, "#7A7A7A"),
            "expertise": areas,
            "summary": str(row.get("Engagement Summary") or "No engagement summary has been entered yet.").strip(),
        }

    links = []
    expertise_to_orgs = {}
    for row in link_rows:
        expertise = str(row.get("Expertise") or "").strip()
        org = str(row.get("Organization") or "").strip()
        if not expertise or not org or org not in organizations:
            continue
        links.append({"expertise": expertise, "organization": org})
        expertise_to_orgs.setdefault(expertise, []).append(org)

    expertise = []
    for i, row in enumerate(expertise_rows, start=1):
        name = str(row.get("Expertise") or "").strip()
        if not name:
            continue
        try:
            order = int(row.get("Display Order") or i)
        except (TypeError, ValueError):
            order = i
        org_names = sorted(set(expertise_to_orgs.get(name, [])))
        expertise.append(
            {
                "name": name,
                "order": order,
                "description": str(row.get("Description") or "").strip(),
                "organization_count": len(org_names),
                "total_engagement": sum(organizations[o]["score"] for o in org_names),
            }
        )

    expertise.sort(key=lambda x: (x["order"], x["name"]))
    org_list = sorted(organizations.values(), key=lambda x: (-x["score"], x["name"]))

    observed_types = {o["type"] for o in org_list}
    types = [t for t in TYPE_ORDER if t in observed_types]
    types.extend(sorted(observed_types.difference(types)))
    type_counts = {t: sum(1 for o in org_list if o["type"] == t) for t in types}

    return {
        "center": "AUiX",
        "expertise": expertise,
        "organizations": org_list,
        "links": links,
        "types": types,
        "type_counts": type_counts,
        "stats": {
            "organizations": len(org_list),
            "expertise": len(expertise),
            "relationships": len(links),
            "total_engagement": int(sum(o["score"] for o in org_list)),
        },
    }


def build_component(network: dict) -> str:
    payload = json.dumps(network, ensure_ascii=False).replace("</", "<\\/")
    template = r'''
<style>
  :root {
    --bg: #F7F8FA;
    --panel: #FFFFFF;
    --text: #17212B;
    --muted: #667085;
    --border: #D8DEE6;
    --auix: #172A46;
    --line: #C7CED8;
  }
  * { box-sizing: border-box; }
  .auix-app { font-family: Inter, ui-sans-serif, system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif; color: var(--text); }
  .view-tabs { display:flex; gap:7px; margin:0 0 12px; flex-wrap:wrap; }
  .view-btn {
    border:1px solid var(--border); background:var(--panel); color:var(--text); border-radius:12px;
    padding:9px 14px; font-size:13px; font-weight:700; cursor:pointer; transition:all .15s ease;
  }
  .view-btn.active { background:var(--auix); color:white; border-color:var(--auix); }
  .filters { display:flex; gap:8px; flex-wrap:wrap; align-items:center; }
  .filter-btn, .expertise-btn, .org-row, .pin-close {
    border:1px solid var(--border); background:var(--panel); color:var(--text); border-radius:999px;
    padding:7px 10px; font-size:12px; cursor:pointer; transition:all .15s ease;
  }
  .filter-btn:hover, .expertise-btn:hover { border-color:#AAB4C2; }
  .filter-btn.active { box-shadow: inset 0 0 0 2px currentColor; }
  .expertise-btn.active { color:#FFFFFF; border-color:transparent; font-weight:700; }
  .expertise-btn:disabled { opacity:.34; cursor:not-allowed; }
  .expertise-strip { display:flex; gap:7px; flex-wrap:wrap; margin:10px 0 7px; }
  .selection-key { display:flex; gap:8px; flex-wrap:wrap; min-height:0; align-items:center; margin:0 0 10px; }
  .selection-chip { display:inline-flex; align-items:center; gap:6px; border-radius:999px; padding:4px 8px; color:#FFFFFF; font-size:11px; font-weight:700; }
  .selection-help { color:var(--muted); font-size:11px; }
  .hidden { display:none !important; }
  .layout { display:grid; grid-template-columns:minmax(0, 2fr) minmax(330px, .88fr); gap:14px; align-items:start; }
  .network-wrap { position:relative; border:1px solid var(--border); border-radius:18px; background:var(--bg); overflow:hidden; }
  #network { width:100%; height:auto; display:block; aspect-ratio: 1.12 / 1; }
  .side { border:1px solid var(--border); border-radius:18px; background:var(--panel); padding:14px; min-height:760px; }
  .side-head { display:flex; justify-content:space-between; align-items:flex-start; gap:8px; }
  .side h3 { margin:0 0 4px; font-size:17px; }
  .side .sub { color:var(--muted); font-size:12px; margin-bottom:10px; }
  .pin-count { font-size:11px; color:var(--muted); white-space:nowrap; }
  .pinned-grid { display:flex; flex-direction:column; gap:8px; margin:8px 0 14px; }
  .pin-card { border:1px solid var(--border); border-radius:12px; padding:10px; background:#FBFCFD; }
  .pin-card.current { border-color:var(--auix); box-shadow: inset 3px 0 0 var(--auix); }
  .pin-card-head { display:flex; align-items:flex-start; justify-content:space-between; gap:8px; }
  .pin-name { font-weight:700; font-size:13px; line-height:1.25; }
  .pin-close { border:0; padding:2px 6px; font-size:14px; background:transparent; color:var(--muted); }
  .pin-meta { display:flex; gap:5px; flex-wrap:wrap; margin-top:5px; }
  .pin-summary { font-size:11px; line-height:1.45; margin:7px 0; color:#344054; }
  .search { width:100%; border:1px solid var(--border); border-radius:10px; padding:9px 10px; font-size:13px; margin:3px 0 9px; }
  .directory-title { font-size:12px; font-weight:700; margin-top:3px; }
  .org-list { display:flex; flex-direction:column; gap:6px; max-height:260px; overflow:auto; padding-right:2px; }
  .org-row { border-radius:10px; text-align:left; padding:8px 9px; display:grid; grid-template-columns:10px 1fr auto; align-items:center; gap:8px; }
  .org-row:hover { background:#F8FAFC; }
  .org-row.active { border-color:var(--auix); background:#F5F7FA; }
  .dot { width:9px; height:9px; border-radius:50%; }
  .score { font-variant-numeric:tabular-nums; color:var(--muted); font-size:11px; }
  .badge { display:inline-flex; border-radius:999px; padding:3px 7px; font-size:10px; color:white; }
  .tag { background:#F1F4F7; border-radius:999px; padding:3px 7px; font-size:10px; color:#465260; }
  .match-tag { color:#FFFFFF; }
  .expertise-tags { display:flex; flex-wrap:wrap; gap:4px; }
  .empty { color:var(--muted); font-size:12px; line-height:1.5; padding:8px 0; }
  .svg-label { font-size:11px; font-weight:650; fill:var(--text); pointer-events:none; transition:opacity .15s ease; }
  .svg-small { font-size:9px; fill:var(--muted); pointer-events:none; transition:opacity .15s ease; }
  .svg-node { cursor:pointer; transition:opacity .15s ease; }
  .center-text { fill:white; font-weight:800; font-size:19px; pointer-events:none; }
  .center-sub { fill:#E7ECF3; font-size:9px; pointer-events:none; }
  .org-initials { fill:white; font-size:7.5px; font-weight:800; pointer-events:none; }
  .org-focus-label { font-size:10px; font-weight:700; fill:var(--text); pointer-events:none; paint-order:stroke; stroke:#FFFFFF; stroke-width:3px; stroke-linejoin:round; }
  .org-label-leader { stroke:#98A2B3; stroke-width:.7; opacity:.42; pointer-events:none; }
  .edge { stroke:var(--line); stroke-width:1.1; opacity:.58; transition:opacity .15s ease; }
  @media (max-width: 780px) {
    .view-tabs { flex-wrap:nowrap; }
    .view-btn { flex:1 1 50%; }
    .filters, .expertise-strip { flex-wrap:nowrap; overflow-x:auto; padding-bottom:5px; scrollbar-width:thin; }
    .filter-btn, .expertise-btn { flex:0 0 auto; }
    .layout { grid-template-columns:1fr; }
    .network-wrap { order:1; }
    .side { order:2; min-height:0; }
    #network { aspect-ratio:1 / 1; }
    .pinned-grid { display:grid; grid-auto-flow:column; grid-auto-columns:minmax(260px, 88%); overflow-x:auto; gap:8px; padding-bottom:7px; scroll-snap-type:x proximity; }
    .pin-card { scroll-snap-align:start; }
    .org-list { max-height:340px; }
    .org-focus-label { font-size:9px; }
  }
</style>

<div class="auix-app">
  <div class="view-tabs" aria-label="Network view">
    <button type="button" class="view-btn active" id="typeViewBtn">By Organization Type</button>
    <button type="button" class="view-btn" id="expertiseViewBtn">By Expertise</button>
  </div>

  <div id="expertiseControls" class="hidden">
    <div class="filters" id="typeFilters" aria-label="Organization type filters"></div>
    <div class="expertise-strip" id="expertiseStrip" aria-label="Expertise shortcuts"></div>
    <div class="selection-key" id="selectionKey"></div>
  </div>

  <div class="layout">
    <div class="network-wrap">
      <svg id="network" viewBox="0 0 900 800" role="img" aria-label="AUiX stakeholder network"></svg>
    </div>
    <aside class="side" aria-live="polite">
      <div class="side-head">
        <div>
          <h3 id="sideTitle">Pinned organizations</h3>
          <div class="sub" id="sideSub">Click any organization node to build a comparison set.</div>
        </div>
        <div class="pin-count" id="pinCount">0 / 4 pinned</div>
      </div>
      <div class="pinned-grid" id="pinnedGrid">
        <div class="empty">Profiles will stay pinned here. After four are pinned, clicking a new organization replaces the oldest pin.</div>
      </div>
      <div class="directory-title" id="directoryTitle">Organization directory</div>
      <input class="search" id="searchBox" type="search" placeholder="Search organizations" aria-label="Search organizations">
      <div class="org-list" id="orgList"></div>
    </aside>
  </div>
</div>

<script>
(() => {
  const DATA = __NETWORK_DATA__;
  const EXPERTISE_COLORS = ['#E8752E', '#2F6FB0', '#178F7A'];
  const state = {
    view: 'type',
    selectedType: null,
    expertise: [],
    expertiseColors: new Map(),
    organization: null,
    pinned: [],
    types: new Set(DATA.types),
    query: ''
  };

  const svg = document.getElementById('network');
  const strip = document.getElementById('expertiseStrip');
  const filters = document.getElementById('typeFilters');
  const selectionKey = document.getElementById('selectionKey');
  const expertiseControls = document.getElementById('expertiseControls');
  const typeViewBtn = document.getElementById('typeViewBtn');
  const expertiseViewBtn = document.getElementById('expertiseViewBtn');
  const orgList = document.getElementById('orgList');
  const pinnedGrid = document.getElementById('pinnedGrid');
  const pinCount = document.getElementById('pinCount');
  const search = document.getElementById('searchBox');
  const sideSub = document.getElementById('sideSub');
  const directoryTitle = document.getElementById('directoryTitle');

  const orgByName = new Map(DATA.organizations.map(o => [o.name, o]));
  const linkSet = new Set(DATA.links.map(l => `${l.expertise}|||${l.organization}`));
  const fixedOrganizations = [...DATA.organizations].sort((a,b) => a.name.localeCompare(b.name));

  const escapeHtml = (s) => String(s ?? '')
    .replaceAll('&','&amp;').replaceAll('<','&lt;').replaceAll('>','&gt;')
    .replaceAll('"','&quot;').replaceAll("'",'&#039;');

  const initials = (name) => {
    const bits = name.replaceAll('&',' ').replace(/[()]/g, ' ').split(/\s+/).filter(Boolean);
    if (bits.length === 1) return bits[0].slice(0,3).toUpperCase();
    return bits.slice(0,3).map(x => x[0]).join('').toUpperCase();
  };

  function expertiseColor(name) {
    return state.expertiseColors.get(name) || '#E8752E';
  }

  function assignExpertiseColor(name) {
    const used = new Set(state.expertiseColors.values());
    const color = EXPERTISE_COLORS.find(c => !used.has(c)) || EXPERTISE_COLORS[state.expertise.length % EXPERTISE_COLORS.length];
    state.expertiseColors.set(name, color);
  }

  function selectedMatches(org) {
    return state.expertise.filter(ex => linkSet.has(`${ex}|||${org.name}`));
  }

  function passesTypeFilter(org) {
    return state.types.has(org.type);
  }

  function passesSearch(org) {
    const q = state.query.trim().toLowerCase();
    return !q || org.name.toLowerCase().includes(q) || org.summary.toLowerCase().includes(q) || org.expertise.some(x => x.toLowerCase().includes(q)) || org.type.toLowerCase().includes(q);
  }

  function directoryOrganizations() {
    const searching = state.query.trim().length > 0;
    if (searching) return fixedOrganizations.filter(org => passesSearch(org));

    if (state.view === 'type') {
      return fixedOrganizations.filter(org => !state.selectedType || org.type === state.selectedType);
    }

    return fixedOrganizations.filter(org => {
      const matches = selectedMatches(org);
      return passesTypeFilter(org) && (!state.expertise.length || matches.length > 0);
    });
  }

  function setView(view) {
    state.view = view;
    typeViewBtn.classList.toggle('active', view === 'type');
    expertiseViewBtn.classList.toggle('active', view === 'expertise');
    expertiseControls.classList.toggle('hidden', view !== 'expertise');
    renderAll();
  }

  function renderControls() {
    typeViewBtn.classList.toggle('active', state.view === 'type');
    expertiseViewBtn.classList.toggle('active', state.view === 'expertise');
    expertiseControls.classList.toggle('hidden', state.view !== 'expertise');

    filters.innerHTML = '';
    DATA.types.forEach(type => {
      const btn = document.createElement('button');
      btn.type = 'button';
      btn.className = 'filter-btn' + (state.types.has(type) ? ' active' : '');
      btn.textContent = type;
      btn.style.color = DATA.organizations.find(o => o.type === type)?.color || '#667085';
      btn.addEventListener('click', () => {
        if (state.types.has(type)) state.types.delete(type); else state.types.add(type);
        if (state.types.size === 0) state.types.add(type);
        renderAll();
      });
      filters.appendChild(btn);
    });

    strip.innerHTML = '';
    DATA.expertise.forEach(ex => {
      const selected = state.expertise.includes(ex.name);
      const btn = document.createElement('button');
      btn.type = 'button';
      btn.className = 'expertise-btn' + (selected ? ' active' : '');
      btn.textContent = `${ex.name} (${ex.organization_count})`;
      btn.disabled = !selected && state.expertise.length >= 3;
      if (selected) btn.style.background = expertiseColor(ex.name);
      btn.title = btn.disabled ? 'Maximum of 3 expertise areas selected. Deselect one first.' : ex.description;
      btn.addEventListener('click', () => selectExpertise(ex.name));
      strip.appendChild(btn);
    });

    selectionKey.innerHTML = '';
    if (state.expertise.length) {
      state.expertise.forEach(ex => {
        const chip = document.createElement('span');
        chip.className = 'selection-chip';
        chip.style.background = expertiseColor(ex);
        chip.textContent = ex;
        selectionKey.appendChild(chip);
      });
      const help = document.createElement('span');
      help.className = 'selection-help';
      help.textContent = `${state.expertise.length} / 3 selected`;
      selectionKey.appendChild(help);
    }
  }

  function selectExpertise(name) {
    const index = state.expertise.indexOf(name);
    if (index >= 0) {
      state.expertise.splice(index, 1);
      state.expertiseColors.delete(name);
    } else if (state.expertise.length < 3) {
      assignExpertiseColor(name);
      state.expertise.push(name);
    }
    state.organization = null;
    renderAll();
  }

  function selectType(name) {
    state.selectedType = state.selectedType === name ? null : name;
    state.organization = null;
    renderAll();
  }

  function pinOrganization(name) {
    const existing = state.pinned.indexOf(name);
    if (existing >= 0) state.pinned.splice(existing, 1);
    else if (state.pinned.length >= 4) state.pinned.shift();
    state.pinned.push(name);
  }

  function selectOrganization(name) {
    state.organization = name;
    pinOrganization(name);
    renderAll();
  }

  function removePin(name) {
    state.pinned = state.pinned.filter(x => x !== name);
    if (state.organization === name) state.organization = state.pinned[state.pinned.length - 1] || null;
    renderAll();
  }

  function svgEl(tag, attrs={}) {
    const el = document.createElementNS('http://www.w3.org/2000/svg', tag);
    Object.entries(attrs).forEach(([k,v]) => el.setAttribute(k, v));
    return el;
  }

  function addText(x, y, text, cls, anchor='middle', opacity=1, fill=null) {
    const attrs = {x, y, 'text-anchor': anchor, class: cls, opacity};
    if (fill) attrs.fill = fill;
    const t = svgEl('text', attrs);
    t.textContent = text;
    svg.appendChild(t);
    return t;
  }

  function fixedOrgPositions(cx, cy, orgR) {
    const positions = new Map();
    fixedOrganizations.forEach((org, i) => {
      const angle = -Math.PI/2 + (i / fixedOrganizations.length) * Math.PI * 2;
      positions.set(org.name, {x: cx + orgR*Math.cos(angle), y: cy + orgR*Math.sin(angle), angle});
    });
    return positions;
  }

  function groupedTypeLayout(cx, cy, typeR, orgR) {
    const orgPositions = new Map();
    const typePositions = new Map();
    const ordered = [];

    // Keep organizations of the same type adjacent, but preserve one perfectly
    // even angular step around the full outer ring.
    DATA.types.forEach(type => {
      fixedOrganizations
        .filter(org => org.type === type)
        .sort((a, b) => a.name.localeCompare(b.name))
        .forEach(org => ordered.push(org));
    });

    const total = Math.max(1, ordered.length);
    const step = (Math.PI * 2) / total;
    const start = -Math.PI / 2;

    ordered.forEach((org, i) => {
      const angle = start + i * step;
      orgPositions.set(org.name, {
        x: cx + orgR * Math.cos(angle),
        y: cy + orgR * Math.sin(angle),
        angle
      });
    });

    // Put each type node at the midpoint of its organization's arc, so the
    // inner node sits directly inside the group it represents.
    let offset = 0;
    DATA.types.forEach(type => {
      const count = ordered.filter(org => org.type === type).length;
      if (!count) return;
      const midpointIndex = offset + (count - 1) / 2;
      const angle = start + midpointIndex * step;
      typePositions.set(type, {
        x: cx + typeR * Math.cos(angle),
        y: cy + typeR * Math.sin(angle),
        angle
      });
      offset += count;
    });

    return {orgPositions, typePositions};
  }

  function drawOrganization(org, p, cx, cy, orgR, opacity, rings, showFullLabel) {
    if (opacity <= 0) return;
    const current = state.organization === org.name;
    const pinned = state.pinned.includes(org.name);
    const baseRadius = Math.max(10, Math.min(17, 9 + Math.sqrt(Math.max(0, org.score)) * 1.15));
    const radius = current ? baseRadius + 3 : pinned ? baseRadius + 1.5 : baseRadius;
    const g = svgEl('g', {class:'svg-node', opacity});
    const c = svgEl('circle', {
      cx:p.x, cy:p.y, r:radius,
      fill:org.color,
      stroke: current ? '#172A46' : pinned ? '#344054' : '#FFFFFF',
      'stroke-width': current ? 4 : pinned ? 3 : 2
    });
    const title = svgEl('title');
    title.textContent = `${org.name} · ${org.type} · Engagement ${org.score} · ${org.expertise.join(', ')}`;
    c.appendChild(title);
    g.appendChild(c);
    const txt = svgEl('text', {x:p.x, y:p.y+2.7, 'text-anchor':'middle', class:'org-initials'});
    txt.textContent = initials(org.name);
    g.appendChild(txt);
    g.addEventListener('click', () => selectOrganization(org.name));
    svg.appendChild(g);

    rings.forEach((ring, ringIndex) => {
      const outline = svgEl('circle', {
        cx:p.x, cy:p.y,
        r:radius + 4 + ringIndex*4,
        fill:'none', stroke:ring.color,
        'stroke-width':'2.2', opacity:'0.95', 'pointer-events':'none'
      });
      svg.appendChild(outline);
    });

    if (showFullLabel) {
      const labelR = orgR + 30;
      const lx = cx + labelR*Math.cos(p.angle);
      const ly = cy + labelR*Math.sin(p.angle);
      const cang = Math.cos(p.angle);
      const anchor = cang > 0.18 ? 'start' : cang < -0.18 ? 'end' : 'middle';
      const leaderStartR = orgR + radius + 2 + rings.length*4;
      const leaderEndR = orgR + 24;
      const leader = svgEl('line', {
        x1: cx + leaderStartR*Math.cos(p.angle),
        y1: cy + leaderStartR*Math.sin(p.angle),
        x2: cx + leaderEndR*Math.cos(p.angle),
        y2: cy + leaderEndR*Math.sin(p.angle),
        class:'org-label-leader'
      });
      svg.appendChild(leader);
      addText(lx, ly+3, org.name, 'org-focus-label', anchor, 1);
    }
  }

  function renderTypeNetwork() {
    svg.innerHTML = '';
    const cx = 450, cy = 390;
    const typeR = 176;
    const orgR = 326;
    const hasSelection = Boolean(state.selectedType);
    const groupedLayout = groupedTypeLayout(cx, cy, typeR, orgR);
    const typePositions = groupedLayout.typePositions;
    const orgPositions = groupedLayout.orgPositions;

    DATA.types.forEach(type => {
      const p = typePositions.get(type);
      const selected = state.selectedType === type;
      const opacity = hasSelection ? (selected ? 0.9 : 0.06) : 0.62;
      const line = svgEl('line', {x1:cx, y1:cy, x2:p.x, y2:p.y, stroke:DATA.organizations.find(o => o.type === type)?.color || '#7A7A7A', 'stroke-width':selected ? 2.5 : 1.2, opacity});
      svg.appendChild(line);
    });

    DATA.organizations.forEach(org => {
      if (hasSelection && org.type !== state.selectedType) return;
      const tp = typePositions.get(org.type);
      const op = orgPositions.get(org.name);
      if (!tp || !op) return;
      const edge = svgEl('line', {
        x1:tp.x, y1:tp.y, x2:op.x, y2:op.y,
        stroke:org.color, 'stroke-width':hasSelection ? '1.55' : '1.05', opacity:hasSelection ? '0.36' : '0.16'
      });
      svg.appendChild(edge);
    });

    fixedOrganizations.forEach(org => {
      const p = orgPositions.get(org.name);
      const typeMatch = !hasSelection || org.type === state.selectedType;
      const searchMatch = passesSearch(org);
      const current = state.organization === org.name;
      const pinned = state.pinned.includes(org.name);

      let opacity = hasSelection ? (typeMatch ? 1 : 0.055) : 0.64;
      if (state.query && searchMatch) opacity = 1;
      else if (state.query && !searchMatch) opacity = 0.04;

      const rings = hasSelection && typeMatch ? [{color:org.color}] : [];
      const showFullLabel = opacity > 0.2 && ((state.query && searchMatch) || (hasSelection && typeMatch) || current || pinned);
      drawOrganization(org, p, cx, cy, orgR, opacity, rings, showFullLabel);
    });

    DATA.types.forEach(type => {
      const p = typePositions.get(type);
      const selected = state.selectedType === type;
      const dimmed = hasSelection && !selected;
      const opacity = dimmed ? 0.07 : 1;
      const color = DATA.organizations.find(o => o.type === type)?.color || '#7A7A7A';
      const g = svgEl('g', {class:'svg-node', opacity});
      const circle = svgEl('circle', {
        cx:p.x, cy:p.y, r:selected ? 38 : 32,
        fill:selected ? color : '#FFFFFF',
        stroke:color, 'stroke-width':selected ? 4 : 3
      });
      const title = svgEl('title');
      title.textContent = `${type}: ${DATA.type_counts[type] || 0} organizations`;
      circle.appendChild(title);
      g.appendChild(circle);
      g.addEventListener('click', () => selectType(type));
      svg.appendChild(g);

      const dx = p.x - cx, dy = p.y - cy;
      const len = Math.max(1, Math.sqrt(dx*dx + dy*dy));
      const lx = p.x + (dx/len)*50;
      const ly = p.y + (dy/len)*50;
      let anchor = 'middle';
      if (dx > 70) anchor = 'start';
      if (dx < -70) anchor = 'end';
      addText(lx, ly, type, 'svg-label', anchor, dimmed ? 0.07 : 1, selected ? color : null);
      addText(lx, ly+13, `${DATA.type_counts[type] || 0} orgs`, 'svg-small', anchor, dimmed ? 0.07 : 1);
    });

    const center = svgEl('circle', {cx, cy, r:57, fill:'#172A46', stroke:'#FFFFFF', 'stroke-width':5});
    svg.appendChild(center);
    addText(cx, cy+1, DATA.center, 'center-text');
    addText(cx, cy+18, hasSelection ? state.selectedType : 'stakeholder ecosystem', 'center-sub');
  }

  function renderExpertiseNetwork() {
    svg.innerHTML = '';
    const cx = 450, cy = 390;
    const exR = 176;
    const orgR = 326;
    const expertisePositions = new Map();
    const orgPositions = fixedOrgPositions(cx, cy, orgR);
    const hasSelection = state.expertise.length > 0;

    DATA.expertise.forEach((ex, i) => {
      const angle = -Math.PI/2 + (i / DATA.expertise.length) * Math.PI * 2;
      expertisePositions.set(ex.name, {x: cx + exR*Math.cos(angle), y: cy + exR*Math.sin(angle), angle});
    });

    DATA.expertise.forEach(ex => {
      const p = expertisePositions.get(ex.name);
      const selected = state.expertise.includes(ex.name);
      const opacity = hasSelection ? (selected ? 0.88 : 0.05) : 0.52;
      const line = svgEl('line', {x1:cx, y1:cy, x2:p.x, y2:p.y, class:'edge', opacity});
      if (selected) {
        line.setAttribute('stroke', expertiseColor(ex.name));
        line.setAttribute('stroke-width', '2');
      }
      svg.appendChild(line);
    });

    state.expertise.forEach(exName => {
      const ep = expertisePositions.get(exName);
      const color = expertiseColor(exName);
      fixedOrganizations.forEach(org => {
        if (!passesTypeFilter(org) || !linkSet.has(`${exName}|||${org.name}`)) return;
        const op = orgPositions.get(org.name);
        const edge = svgEl('line', {
          x1:ep.x, y1:ep.y, x2:op.x, y2:op.y,
          stroke:color, 'stroke-width':'1.55', opacity:'0.34'
        });
        svg.appendChild(edge);
      });
    });

    fixedOrganizations.forEach(org => {
      const p = orgPositions.get(org.name);
      const typeActive = passesTypeFilter(org);
      const searchMatch = passesSearch(org);
      const matches = selectedMatches(org);
      const current = state.organization === org.name;
      const pinned = state.pinned.includes(org.name);

      let opacity = 0.60;
      if (!typeActive) opacity = 0.035;
      else if (state.query && searchMatch) opacity = 1;
      else if (state.query && !searchMatch) opacity = 0.04;
      else if (hasSelection && matches.length === 0) opacity = 0.055;
      else if (hasSelection) opacity = 1;

      const rings = matches.map(exName => ({color:expertiseColor(exName)}));
      const showFullLabel = opacity > 0.2 && ((state.query && searchMatch && typeActive) || (hasSelection && matches.length > 0 && typeActive) || current || pinned);
      drawOrganization(org, p, cx, cy, orgR, opacity, rings, showFullLabel);
    });

    DATA.expertise.forEach(ex => {
      const p = expertisePositions.get(ex.name);
      const selected = state.expertise.includes(ex.name);
      const dimmed = hasSelection && !selected;
      const opacity = dimmed ? 0.065 : 1;
      const color = selected ? expertiseColor(ex.name) : '#E8752E';
      const g = svgEl('g', {class:'svg-node', opacity});
      const circle = svgEl('circle', {
        cx:p.x, cy:p.y, r:selected ? 33 : 26,
        fill:selected ? color : '#FFFFFF',
        stroke:color, 'stroke-width':selected ? 4 : 2
      });
      const title = svgEl('title');
      title.textContent = `${ex.name}: ${ex.description}`;
      circle.appendChild(title);
      g.appendChild(circle);
      g.addEventListener('click', () => selectExpertise(ex.name));
      svg.appendChild(g);

      const dx = p.x - cx, dy = p.y - cy;
      const len = Math.max(1, Math.sqrt(dx*dx + dy*dy));
      const lx = p.x + (dx/len)*42;
      const ly = p.y + (dy/len)*42;
      let anchor = 'middle';
      if (dx > 70) anchor = 'start';
      if (dx < -70) anchor = 'end';
      const labelOpacity = dimmed ? 0.055 : 1;
      addText(lx, ly, ex.name, 'svg-label', anchor, labelOpacity, selected ? color : null);
      addText(lx, ly+13, `${ex.organization_count} orgs`, 'svg-small', anchor, labelOpacity);
    });

    const center = svgEl('circle', {cx, cy, r:57, fill:'#172A46', stroke:'#FFFFFF', 'stroke-width':5});
    svg.appendChild(center);
    addText(cx, cy+1, DATA.center, 'center-text');
    const centerSub = !hasSelection ? 'expertise ecosystem' : state.expertise.length === 1 ? state.expertise[0] : `${state.expertise.length} expertise areas`;
    addText(cx, cy+18, centerSub, 'center-sub');
  }

  function renderNetwork() {
    if (state.view === 'type') renderTypeNetwork();
    else renderExpertiseNetwork();
  }

  function renderPinned() {
    pinCount.textContent = `${state.pinned.length} / 4 pinned`;
    if (!state.pinned.length) {
      pinnedGrid.innerHTML = '<div class="empty">Profiles will stay pinned here. After four are pinned, clicking a new organization replaces the oldest pin.</div>';
      sideSub.textContent = 'Click any organization node to build a comparison set.';
      return;
    }

    sideSub.textContent = 'Up to four profiles stay visible while you switch between views.';
    pinnedGrid.innerHTML = '';
    [...state.pinned].reverse().forEach(name => {
      const org = orgByName.get(name);
      if (!org) return;
      const matches = selectedMatches(org);
      const card = document.createElement('div');
      card.className = 'pin-card' + (state.organization === name ? ' current' : '');
      const matchTags = state.view === 'expertise'
        ? matches.map(ex => `<span class="tag match-tag" style="background:${expertiseColor(ex)}">${escapeHtml(ex)}</span>`).join('')
        : '';
      card.innerHTML = `
        <div class="pin-card-head">
          <div class="pin-name">${escapeHtml(org.name)}</div>
          <button type="button" class="pin-close" aria-label="Remove ${escapeHtml(org.name)} from pinned profiles">×</button>
        </div>
        <div class="pin-meta">
          <span class="badge" style="background:${org.color}">${escapeHtml(org.type)}</span>
          <span class="tag">Engagement ${org.score}</span>
          ${matchTags}
        </div>
        <div class="pin-summary">${escapeHtml(org.summary)}</div>
        <div class="expertise-tags">${org.expertise.map(x => `<span class="tag">${escapeHtml(x)}</span>`).join('')}</div>
      `;
      card.querySelector('.pin-close').addEventListener('click', (e) => {
        e.stopPropagation();
        removePin(name);
      });
      card.addEventListener('click', () => {
        state.organization = name;
        renderAll();
      });
      pinnedGrid.appendChild(card);
    });
  }

  function renderDirectory() {
    const visible = directoryOrganizations();
    if (state.query.trim()) {
      directoryTitle.textContent = `Search results · ${visible.length}`;
    } else if (state.view === 'type' && state.selectedType) {
      directoryTitle.textContent = `${state.selectedType} · ${visible.length} organizations`;
    } else if (state.view === 'expertise' && state.expertise.length) {
      directoryTitle.textContent = `Selected expertise · ${visible.length} matching organizations`;
    } else {
      directoryTitle.textContent = `Organization directory · ${visible.length} visible`;
    }

    orgList.innerHTML = '';
    if (!visible.length) {
      orgList.innerHTML = '<div class="empty">No organizations match the current view or search.</div>';
      return;
    }
    visible.forEach(org => {
      const btn = document.createElement('button');
      btn.type = 'button';
      btn.className = 'org-row' + (state.organization === org.name ? ' active' : '');
      btn.innerHTML = `<span class="dot" style="background:${org.color}"></span><span>${escapeHtml(org.name)}</span><span class="score">${org.score}</span>`;
      btn.addEventListener('click', () => selectOrganization(org.name));
      orgList.appendChild(btn);
    });
  }

  function resizeFrame() {
    try {
      const h = Math.max(document.documentElement.scrollHeight, document.body.scrollHeight) + 10;
      window.parent.postMessage({isStreamlitMessage:true, type:'streamlit:setFrameHeight', height:h}, '*');
    } catch (e) {}
  }

  function renderAll() {
    renderControls();
    renderNetwork();
    renderPinned();
    renderDirectory();
    window.setTimeout(resizeFrame, 30);
  }

  typeViewBtn.addEventListener('click', () => setView('type'));
  expertiseViewBtn.addEventListener('click', () => setView('expertise'));

  search.addEventListener('input', (e) => {
    state.query = e.target.value || '';
    renderNetwork();
    renderDirectory();
    window.setTimeout(resizeFrame, 30);
  });

  window.addEventListener('resize', () => window.setTimeout(resizeFrame, 80));
  renderAll();
})();
</script>
'''
    return template.replace("__NETWORK_DATA__", payload)


st.set_page_config(
    page_title="AUiX Network",
    page_icon="🕸️",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.title("AUiX Network")
with st.sidebar:
    st.header("Data")
    st.write("The app ships with the cleaned AUiX workbook. Upload an updated workbook using the same sheet structure to refresh both network views without changing the code.")
    uploaded = st.file_uploader("Use another AUiX workbook", type=["xlsx"])
    st.divider()
    st.markdown("**Expected sheets**")
    st.code("Stakeholders\nExpertise_Links\nExpertise")

try:
    network = load_network_data(uploaded.getvalue() if uploaded else DEFAULT_DATA.read_bytes())
except Exception as exc:
    st.error(f"Could not load the workbook: {exc}")
    st.stop()

stats = network["stats"]
c1, c2, c3, c4 = st.columns(4)
c1.metric("Organizations", stats["organizations"])
c2.metric("Expertise areas", stats["expertise"])
c3.metric("Expertise relationships", stats["relationships"])
c4.metric("Total engagement", stats["total_engagement"])

components.html(build_component(network), height=1180, scrolling=False)
